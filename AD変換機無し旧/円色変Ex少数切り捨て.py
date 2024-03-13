import pygame
from pygame.locals import *
import openpyxl
import numpy as np
import math
import time
import datetime as dt

#保存するExcelファイルの名前指定をしてから実行
print("\n DAY")
Day =("{}".format(input()))

print("\n Title")
T_name =("T{}".format(input()))

print(Day,T_name)

name  =("{}_{}.xlsx".format(Day,T_name))

start =("{}".format(input()))



#pygameの初期化
pygame.init()
#１秒間に対するフレーム数
FPS = 25
#1秒間に対するサンプル数
SPS = 2000

#表示倍率変数a アンプまたはズーム
a = 3.6
#周期T
T = 4
#準備時間
pre = 10
#制限時間
limit = 60 + pre

#原点
x0,y0 = 900,100
#アームの長さ
L =400
yy = y0 + L
#周期運動するボールの初期位置
XB0, YB0 = x0, yy

#Excelファイルを読み込む
book = openpyxl.load_workbook("円運動テンプレート.xlsx")
sheet = book.active
sheet.tittle = "round"
i = 2
sheet.cell(row = i, column = 1).value = dt.datetime.now()


#スクリーンの画面サイズのオブジェクト
screen = pygame.display.set_mode((1880, 1050))
#時間管理のオブジェクト
clock = pygame.time.Clock()
#フォントのオブジェクト
font = pygame.font.Font(None, 50)
#色のオブジェクト
BLUE, RED, GREEN = ((0, 0, 255), (255, 0, 0), (0, 255, 0))
WHITE, GLAY = ((255, 255, 255), (160, 160, 160))
PINK,ORANGE,LIGHTBLUE= ((255,0,255),(255,128,0),(51,153,255))
# マウスカーソル非表示
# pygame.mouse.set_visible(False)
# マウスカーソル初期位置
pygame.mouse.set_pos((x0, y0))

#円でかさ(,マウスピンク)
#Xa_BALL , Ya_BALL = int(xc) , int(yc)
r1 , r2 = 20*a , 7*a
#count = 0
#judge = False

#pygameの初期化
pygame.init()
LOOP = True
while LOOP:
#for row in range(400):
    for event in pygame.event.get():
        if event.type == pygame.QUIT:
            book.save("{}".format(name))
            LOOP = False    #ウインドウを閉じたら、その時点で記録して終了
            
    TIME_S = time.time()            
    #clockオブジェクトを更新する
    clock.tick(FPS)
    smp = clock.get_time()
    #print(smp)
    
    #V仮定(動かすために適当書く)
    #V0=2
    #V1=4
    #もしくは最初から角度決めちゃう
    deg1 = 30
    deg2 = 150
    #角度計算
    #deg1 = 182.15 - (340/3.3) * V0
    #deg2 = 357.97 - (340/3.3) * V1
    
    #関節と手先の座標位置
    x1 = x0 + L * np.cos(math.radians(deg1))
    y1 = y0 + L * np.sin(math.radians(deg1))
    x2 = x1 + L * np.cos(math.radians(deg2))
    y2 = y1 + L * np.sin(math.radians(deg2))
    
    #周期運動の定義
    #周期
    T = 10
    
    #pygameを初期化してからこれまでの時間取得
    t = pygame.time.get_ticks() / 1000
    
    #角振動数
    omega = 2 * math.pi / T #2*3.14/T
    ##周期運動するボールの位置
    X_BALL = XB0 + 450 * np.sin(omega * t)
    Y_BALL = YB0 + 450 * np.sin(omega * t + math.pi / 2 )
    
    #マウス位置
    for event in pygame.event.get():
        if event.type == MOUSEMOTION:
            xm,ym = event.pos
            
    xm,ym = event.pos
    
    mBd = math.sqrt((X_BALL -xm) ** 2 + (Y_BALL - ym)**2)
    
   #ウインドウに制限時間を表示
    sur = font.render(str("TIME"),True,WHITE)
    screen.blit(sur,[30,350])
    if (pre-t) >= 0:  #準備時間内
        sur = font.render(str(f'{pre-t:.2f}'),True,RED)
        screen.blit(sur,[180,350])
    elif (limit -t) >= 0:  #制限時間内
        sur = font.render(str(f'{limit-t:.2f}'),True,WHITE)
        screen.blit(sur,[180,350])
    elif (limit -t) < 0:  #制限時間外
        if (limit + 5) - t >= 0 : #制限時間＋5秒内
            sur = font.render(str("END"),True,RED)
            screen.blit(sur,[180,350])
        elif (limit + 5) - t < 0 :  #制限時間＋5秒経過
            book.save(name)
            LOOP = False    #実行時間が制限時間以上になると保存して終了
    
    #スクリーンに表示
    pygame.draw.circle(screen,LIGHTBLUE, (int(x0), int(yy)),450,2)
    pygame.draw.rect(screen, WHITE, (int(x0), int(yy), 16, 16))
    #pygame.draw.rect(screen, WHITE, (int(x0), int(y0), 16, 16))

    #pygame.draw.rect(screen, RED, (int(x2), int(y2), 16, 16))
    
    #十字線
    pygame.draw.line(screen,LIGHTBLUE,(int(x0),int(50)),(int(x0),int(950)),1)
    
    #pygame.draw.line(screen, LIGHTBLUE , ( int(xc)        , int(yc-(A2*a)) ) , ( int(xc)        , int(yc+(A2*a)) ), int(2) )
    
    if mBd > r1:
            pygame.draw.circle(screen,RED, (int(X_BALL), int(Y_BALL)),int(r1),12)      
    else:
            pygame.draw.circle(screen, GREEN, (int(X_BALL), int(Y_BALL)),int(r1),12)
        
    #マウスの表示(ピンク円)
    pygame.draw.circle(screen , (PINK)  , ( int(xm) , int(ym) ) , int(r2))
    
    text = font.render(f'{math.floor(mBd)}', True, (BLUE))    
    screen.blit(text, [0, 0])
        
    #画面を更新する
    pygame.display.update()
    #画面を真っ白にする
    screen.fill((0, 0, 0))
    
    #Excelのセルに入力
    sheet.cell(row = i, column = 2).value = t
    sheet.cell(row = i, column = 3).value = (math.floor(X_BALL)-x0)
    sheet.cell(row = i, column = 4).value = (math.floor(Y_BALL)-yy)
    sheet.cell(row = i, column = 5).value = xm-x0
    sheet.cell(row = i, column = 6).value = ym-yy
    sheet.cell(row = i, column = 7).value = (math.floor(mBd))
    i += 1
    
    book.save(name)

pygame.quit()